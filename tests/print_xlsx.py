import openpyxl
import pandas as pd 
import os
from tkinter import filedialog
import tkinter as tk

def load_coverage_from_txt(txt_file_path):

    subject = txt_file_path.split('\\')[6]
    leg = txt_file_path.split('\\')[-1].split('_')[2][0]

    try:
        with open(txt_file_path, 'r') as file:
            lines = file.readlines()
            for line in lines:
                if line.startswith('Normalized Area Covered: '):
                    coverage_value = float(line.split(':')[1].strip()[:-1])
                    print(f"Normalized Area Covered {subject}_{leg}: {coverage_value} %")
    except FileNotFoundError:
        print(f"File {txt_file_path} does not exist")
        coverage_value = None
        
    return coverage_value

def load_workbook_to_df(filename, sheet_name):
    # Load the existing XLSX file
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    column_names = [cell.value for cell in sheet[1]]
    data = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
    df = pd.DataFrame(data, columns=column_names)
    return df, workbook

def add_df_to_xlsx(df, filename, sheet_name):

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    column_names = [cell.value for cell in sheet[1]]
    data = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
    df2 = pd.DataFrame(data, columns=column_names)

    # Clear the existing data in the sheet (optional)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None  # This clears existing data

    # Write the DataFrame headers to the first row
    for col, header in enumerate(df.columns):
        sheet.cell(row=1, column=col+1).value = header

    # Write the DataFrame data to the existing sheet
    for idx, row in df.iterrows():
        for col, value in enumerate(row):
            sheet.cell(row=idx+2, column=col+1).value = value

    return workbook

def popup_tickbox(list_names):
    root = tk.Tk()
    root.title("Select Subjects")

    # Create a list to store the checkbox variables
    checkboxes = []
    columns = 3
    selected_names = [] 

    # Create a checkbox for each name in the list
    for i, name in enumerate(list_names):
        var = tk.BooleanVar()
        checkbox = tk.Checkbutton(root, text=name, variable=var)
        checkbox.grid(row=i // columns, column=i % columns)
        checkboxes.append(var)

    # Get selected names
    def get_selected_and_close():
        for i, checkbox in enumerate(checkboxes):
            if checkbox.get():
                selected_names.append(list_names[i])
        
        # Close the window
        root.destroy()

    # Create and place the OK button
    ok_button = tk.Button(root, text="OK", command=get_selected_and_close)
    ok_button.grid(row=len(list_names) // columns + 1, column=1, columnspan=2, sticky=tk.W + tk.E)

    root.mainloop()

    return selected_names  # Return the list of selected names
       
def add_coverages_to_xlsx(path_coverage='', filename='',sheet_name=''):
    
    if not path_coverage or not os.path.exists(path_coverage):
        path_coverage = filedialog.askdirectory(title='Select folder of coverage')
        if not os.path.exists(path_coverage):
            print(f"Path {path_coverage} does not exist")
            return
    
    if not filename or not os.path.exists(filename):
        filename = filedialog.askopenfilename(title='Select XLSX file', filetypes=[('Excel files', '*.xlsx')])
    
    if not sheet_name:
        workbook = openpyxl.load_workbook(filename)
        sheet_names = workbook.sheetnames
        sheet_name = popup_tickbox(sheet_names)
        sheet_name = sheet_name[0]
        print(f"Sheet name: {sheet_name}")

    # Load the existing XLSX file
    df, workbook = load_workbook_to_df(filename, sheet_name)

    column_name_l = 'L_acetabular coverage_BG'
    column_name_r = 'R_acetabular coverage_BG'

    subjects = [entry for entry in os.listdir(path_coverage) if os.path.isdir(os.path.join(path_coverage, entry))]

    legs = ['l', 'r']
    threshold = '25'  # distance threshold in mm
    print(f"Adding coverage values to {filename}")
    for subject in subjects:
        for leg in legs:
            txt_file_path = f'{path_coverage}\\{subject}\\Meshlab_BG\\femoral_head_{leg}_threshold_{threshold}\\femoral_head_{leg}.txt'
            coverage_value = load_coverage_from_txt(txt_file_path)
            
            if coverage_value is not None and leg == 'l':
                df.loc[df['Subject'] == subject, column_name_l] = coverage_value
            elif leg == 'r':
                df.loc[df['Subject'] == subject, column_name_r] = coverage_value
    
    workbook = add_df_to_xlsx(df, filename, sheet_name)

    # Save the workbook with the same formatting
    workbook.save(filename.replace('.xlsx', '_updated.xlsx'))

if __name__ == "__main__":
    filename = r"C:\Users\Bas\ucloud\MRI_segmentation_BG\ParticipantData and Labelling.xlsx"
    path_coverage = r"C:\Users\Bas\ucloud\MRI_segmentation_BG\acetabular_coverage"
    sheet_name = 'Demographics'
    sheet_name = []
    common_column = 'Subject'
    txt_file_path = r"c:\Users\Bas\ucloud\MRI_segmentation_BG\acetabular_coverage\009\Meshlab_BG\femoral_head_l_threshold_25\femoral_head_l.txt"
    
    add_coverages_to_xlsx(path_coverage, filename, sheet_name)
